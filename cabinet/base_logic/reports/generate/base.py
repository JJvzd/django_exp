import os
from copy import copy

import attr
from openpyxl import Workbook, load_workbook
from openpyxl.styles import colors, PatternFill
from openpyxl.utils import get_column_letter

from settings.settings import MEDIA_ROOT, MEDIA_URL


class ExcelCellData:

    def __init__(self, cell=None, row=None, column=None, value=None, color=None,
                 merge=None, href=None, format=None):
        assert cell or (row and column) or merge
        if cell or merge:
            self.cell = cell
        else:
            self.cell = '%s%s' % (get_column_letter(column), row)
        self.value = value
        self.href = href
        self.color = None
        self.format = format
        if color:
            if hasattr(colors, color):
                self.color = getattr(colors, color)
            else:
                self.color = color
        self.merge = merge

    def __repr__(self):
        return '<ExcelCell %s: %s>' % (self.cell, self.value)


@attr.s(auto_attribs=True)
class BaseReportResult:
    file_path: str
    output_name: str
    error: str = ''


class BaseReport:
    template_name = None
    output_filename = None
    sheets_for_remove = []
    money_format = r'#,##0"р.";[RED]\\-#,##0"р."'
    date_format = r'DD.MM.YYYY'

    def __init__(self):
        self.rows_for_insert = []
        self.cols_for_delete = []

    def get_data(self):
        return {}

    def get_template_name(self):
        return self.template_name

    def get_output_filename(self, extension=None):
        return self.output_filename or 'report.xlsx'

    def add_sheet_for_remove(self, sheet_name):
        self.sheets_for_remove.append(sheet_name)

    def action_insert(self, ws):
        for start, count in self.rows_for_insert:
            for _ in range(count):
                ws.insert_rows(start)
                for col in range(1, 50):
                    ws.cell(start, col)._style = copy(ws.cell(start + 1, col)._style)
                    ws.cell(start, col).number_format = ws.cell(start + 1,
                                                                col).number_format
                    ws.row_dimensions[start + _ + 1] = ws.row_dimensions[start]

    def delete_col(self, col):
        self.cols_for_delete.append(col)

    def add_insert_rows(self, start, count):
        self.rows_for_insert.append((start, count))

    def write_to_file(self, data, sheets_list=None):
        template_name = self.get_template_name()
        if not template_name:
            wb = Workbook()
            wb.remove_sheet(wb.active)
        else:
            wb = load_workbook(template_name)

        for sheet_name, sheet_data in data.items():
            sheet_parent = None
            if isinstance(sheet_name, tuple):
                sheet_parent = sheet_name[1]
                sheet_name = sheet_name[0]

            if sheet_name not in wb.sheetnames:
                if not sheet_parent:
                    sheet = wb.create_sheet(sheet_name)
                else:
                    sheet = wb.copy_worksheet(wb.get_sheet_by_name(sheet_parent))
                    sheet.title = sheet_name
            else:
                sheet = wb[sheet_name]

            for col in self.cols_for_delete:
                sheet.delete_cols(col)
            self.action_insert(sheet)

            list_merge = []
            for cell_data in sheet_data:
                if cell_data.merge is not None:
                    list_merge.append(cell_data.merge)
                    continue
                if cell_data.value is not None:
                    sheet[cell_data.cell] = cell_data.value
                if cell_data.color is not None:
                    sheet[cell_data.cell].fill = PatternFill(start_color=cell_data.color,
                                                             end_color=cell_data.color,
                                                             fill_type='solid')
                if cell_data.href:
                    sheet[cell_data.cell].hyperlink = cell_data.href

                if cell_data.format:
                    sheet[cell_data.cell].number_format = getattr(
                        self,
                        '%s_format' % cell_data.format
                    )

            for merge in list_merge:
                sheet.merge_cells(merge)

        wb.create_sheet('Empty')
        for sheet_name in self.sheets_for_remove:
            try:
                wb.remove_sheet(wb.get_sheet_by_name(sheet_name))
            except KeyError:
                pass
        if len(wb.sheetnames) > 1:
            wb.remove_sheet(wb.get_sheet_by_name('Empty'))

        name = self.get_output_filename()
        file_path = os.path.join(MEDIA_ROOT, name)
        wb.save(filename=file_path)
        return os.path.join(MEDIA_URL, name)

    def generate_pdf(self):
        path = self.generate().file_path
        path_soffice = os.environ.get('PATH_SOFFICE')
        if path_soffice:
            os.system('%s --headless --convert-to pdf %s' % (path_soffice, path[1:]))
            os.system('mv %s media/' % self.get_output_filename(extension='pdf'))
            return BaseReportResult(
                file_path=os.path.join(r'/media',
                                       self.get_output_filename(extension='pdf')),
                output_name=self.get_output_filename(extension='pdf')
            )
        return BaseReportResult(
            file_path='',
            output_name='',
            error='Не установлен soffice'
        )

    def generate(self):
        data = self.get_data()
        return BaseReportResult(
            file_path=self.write_to_file(data),
            output_name=self.get_output_filename()
        )
