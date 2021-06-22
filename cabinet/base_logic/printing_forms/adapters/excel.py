import os
import re
import shutil
from datetime import datetime

from django.utils import timezone
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment

from cabinet.base_logic.printing_forms.adapters.base import (
    BasePrintFormGenerator, get_temp_path
)
from bank_guarantee.bank_integrations.inbank.print_forms_helper import InbankHelper
from bank_guarantee.bank_integrations.moscombank.print_forms_helper import (
    MoscombankHelper
)
from bank_guarantee.bank_integrations.spb_bank.print_form_helper import SPBHelper
from cabinet.constants.constants import Target
from external_api.helper import get_dict_to_path
from settings.settings import BASE_DIR
from utils.excel import convert_letter_to_index


class ExcelBasePrintFormGenerator(BasePrintFormGenerator):

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.extension = '.xlsx'

    def create_book(self, template_path=None):
        if not template_path and self.template:
            template_path = self.template

        if template_path:
            book = load_workbook(template_path)
        else:
            book = Workbook()

        return book

    def parse_row_and_col(self, cell):
        row = ''
        col = ''
        col_finish = False
        for char in cell:
            if str(char).isnumeric():
                col_finish = True
            if not col_finish:
                col += char
            if col_finish:
                row += char

        return [int(row), int(convert_letter_to_index(col))]

    def convert_letter_to_index(self, column_letter):
        column_letter = column_letter.upper()
        result = 0

        for i in range(len(column_letter)):
            result *= 26
            result += (ord(column_letter[i]) - ord('A') + 1)
        return result

    def convert_index_to_letter(self, column_index):
        column_index -= 1
        numeric = column_index % 26
        letter = chr(65 + numeric)
        num2 = int(column_index / 26)
        if num2 > 0:
            return self.convert_index_to_letter(num2) + letter
        else:
            return letter

    def fill_sheet(self, sheet, data):

        mergedcells = []
        for group in sheet.merged_cells.ranges:
            mergedcells.append(group)

        for group in mergedcells:
            sheet.unmerge_cells(str(group))

        for value in data:
            if len(value) > 3:
                sheet.cell(value[0], value[1]).number_format = value[3]
            sheet.cell(value[0], value[1], value[2])

        for group in mergedcells:
            sheet.merge_cells(str(group))

        return sheet

    def save_book(self, book, output_name=None):
        if output_name:
            path = output_name
        else:
            path = get_temp_path(self.extension)

        book.save(path)
        return path

    def get_empty_file(self):
        path = get_temp_path(self.extension)
        with open(get_temp_path(self.extension), 'wb') as f:
            f.close()
        return path

    def save_by_path(self, path):
        return shutil.copy(path, get_temp_path(self.extension))

    def insert_rows(self, sheet, example_row, data):
        example_row -= 1
        rows = len(data) - 1
        if rows < 1:
            rows = 1
        sheet.insert_rows(example_row + rows, rows)
        for row_offset in range(rows):
            row_id = example_row + row_offset
            for col_id in range(200):

                if example_row == row_id:
                    continue
                sheet.cell(row_id, col_id).value = sheet.cell(example_row, col_id).value
                sheet.cell(row_id, col_id).number_format = sheet.cell(
                    example_row, col_id
                ).number_format

                if sheet.cell(row_id, col_id).value[0] == '=':
                    formula = sheet.cell(row_id, col_id).value
                    pattern = r'([A-Z]+)(%i)' % (example_row + 1)
                    replace = r'\1%i' % (row_id + 1)
                    formula = re.sub(pattern, replace, formula)
                    sheet.cell(row_id, col_id).value = formula

        row_offset = 0
        for row_data in data:
            for col_data in row_data:
                sheet.cell(example_row + row_offset + 1, col_data[0], col_data[1])
            row_offset += 1


class RequestExcelPrintFormGenerator(ExcelBasePrintFormGenerator):
    helper_class = None

    def __init__(self, request, print_form, *args, **kwargs):
        self.request = request
        self.print_form = print_form
        super(RequestExcelPrintFormGenerator, self).__init__(*args, **kwargs)

    def get_helper(self):
        return self.helper_class(
            request=self.request,
            bank=self.request.bank,
        )

    def generate(self):
        return []


class SGBExcelPrintForm(RequestExcelPrintFormGenerator):
    pass


class MetallInvestExcelPrintForm(RequestExcelPrintFormGenerator):
    pass


class MosComBankConclusion(RequestExcelPrintFormGenerator):
    helper_class = MoscombankHelper

    def get_data(self):
        helper = self.get_helper()
        data = {
            'B6': helper.today,
            'H6': helper.quarter_date_end,
            'H7': helper.last_quarter,
            'E9': helper.full_name,
            'E10': helper.profile.reg_inn,
            'E11': helper.profile.reg_state_date,
            'E12': helper.profile.get_tax_system_display(),
            'E15': helper.request.tender.beneficiary_name,
            "E16": helper.request.tender.beneficiary_inn,
            'E17': helper.request.tender.get_federal_law_display(),
            'E18': helper.request.required_amount,
            'F19': helper.request.interval_from,
            'I19': helper.request.interval_to,
            'E20': helper.request.tender.notification_id,
            'E21': (
                    helper.request.prepaid_expense_amount or
                    helper.request.tender.prepayment_amount
            ),
            'E22': helper.request.tender.subject,
            'E23': helper.commission,
            'F27': helper.last_quarter_client.get_value(1100),
            'H27': helper.last_year_quarter_client.get_value(1100),

            'F28': helper.last_quarter_client.get_value(1210),
            'H28': helper.last_year_quarter_client.get_value(1210),

            'F29': helper.last_quarter_client.get_value(1220),
            'H29': helper.last_year_quarter_client.get_value(1220),

            'F30': helper.last_quarter_client.get_value(1230),
            'H30': helper.last_year_quarter_client.get_value(1230),

            'F33': helper.last_quarter_client.get_value(1240),
            'H33': helper.last_year_quarter_client.get_value(1240),

            'F34': helper.last_quarter_client.get_value(1250),
            'H34': helper.last_year_quarter_client.get_value(1250),

            'F35': helper.last_quarter_client.get_value(1200),
            'H35': helper.last_year_quarter_client.get_value(1200),

            'F36': helper.last_quarter_client.get_value(1600),
            'H36': helper.last_year_quarter_client.get_value(1600),

            'F38': helper.last_quarter_client.get_value(1300),
            'H38': helper.last_year_quarter_client.get_value(1300),

            'F39': helper.last_quarter_client.get_value(1410),
            'H39': helper.last_year_quarter_client.get_value(1410),

            'F40': helper.last_quarter_client.get_value(1400),
            'H40': helper.last_year_quarter_client.get_value(1400),

            'F41': helper.last_quarter_client.get_value(1510),
            'H41': helper.last_year_quarter_client.get_value(1510),

            'F42': helper.last_quarter_client.get_value(1520),
            'H42': helper.last_year_quarter_client.get_value(1520),

            'F45': helper.last_quarter_client.get_value(1530),
            'H45': helper.last_year_quarter_client.get_value(1530),

            'F46': helper.last_quarter_client.get_value(1540),
            'H46': helper.last_year_quarter_client.get_value(1540),

            'F47': helper.last_quarter_client.get_value(1550),
            'H47': helper.last_year_quarter_client.get_value(1550),

            'F48': helper.last_quarter_client.get_value(1500),
            'H48': helper.last_year_quarter_client.get_value(1500),

            'F49': helper.last_quarter_client.get_value(1700),
            'H49': helper.last_year_quarter_client.get_value(1700),

            'F51': helper.last_quarter_client.get_value(2110),
            'H51': helper.last_year_quarter_client.get_value(2100),

            'F54': helper.last_quarter_client.get_value(2120),
            'H54': helper.last_year_quarter_client.get_value(2120),

            'F55': helper.last_quarter_client.get_value(2200),
            'H55': helper.last_year_quarter_client.get_value(2200),

            'F56': helper.last_quarter_client.get_value(2330),
            'H56': helper.last_year_quarter_client.get_value(2330),

            'F57': helper.last_quarter_client.get_value(2320),
            'H57': helper.last_year_quarter_client.get_value(2320),

            'F58': helper.last_quarter_client.get_value(2340),
            'H58': helper.last_year_quarter_client.get_value(2340),

            'F60': helper.last_quarter_client.get_value(2300),
            'H60': helper.last_year_quarter_client.get_value(2300),

            'F61': helper.last_quarter_client.get_value(2400),
            'H61': helper.last_year_quarter_client.get_value(2400),

            'F62': helper.last_quarter_client.get_value(5640),
            'H62': helper.last_year_quarter_client.get_value(5640),

            'F63': helper.last_nalog_declaration,
            'H63': helper.last_year_nalog_declaration,

            'F64': helper.last_revenue_scrin,
            'H64': helper.last_year_revenue_scrin,

            'F65': helper.tax_year,
            'E68': helper.sum_court_cases,
            'E69': helper.unfinished_executing_proceeding,
            'E70': helper.arrears_debit,
            'E71': helper.arrears_credit,
            'F72': helper.count_executing_contracts,
            'E73': helper.kind_of_activity,
            'F79': helper.stop_amount_bg,
            'F85': helper.stop_resident_rf,
            'F86': helper.unreliability_egrul,
            'F87': helper.stop_liquidation,
            'F89': helper.stop_many_registration,
            'F91': helper.disqualified_person,
            'F92': helper.inability_to_lead,
            'F93': helper.unscrupulous_supplier,
            'F96': helper.invalid_documents,
            'F97': helper.false_accounting,
            'F98': helper.terror_check,
            'F99': helper.info_193_t,
            'F100': helper.in_list_693_p,
            'F101': helper.criminal_is_gen_dir,
            'F106': helper.is_casino,
            'F107': helper.is_show_business,
            'F108': helper.is_activity_stock,
            'F109': helper.is_sport_business,
            'F110': helper.is_rare_animal_trade,
            'F111': helper.is_prohibited_products,
            'F112': helper.is_credit_company,
            'F113': helper.is_religious_organization,
            'F114': helper.is_building_company,
            'F115': helper.is_rent_property,
            'F116': helper.buy_property,
            'F118': helper.bankrupt,
            'F119': helper.beneficiars_bankrupt,
            'F120': helper.guarantor_bankrupt,
            'F127': helper.negative_credit_history,
            'F128': helper.credit_history_delay_5_day,
            'F129': helper.other_warranty_depreciated,
            'F131': helper.has_unpaid_billing,
            'F133': helper.suspension_decision,
            'F135': helper.debt_fot,
            'F137': helper.has_any_debt,
            'F141': helper.discrepancy_tender_and_request,
        }

        return [[*self.parse_row_and_col(k), v] for k, v in data.items()]

    def generate(self):
        template_path = self.print_form.get_template(self.request, self.request.bank)
        wb = self.create_book(template_path=template_path)
        ws = wb.active
        data = self.get_data()
        self.fill_sheet(ws, data)

        return [self.save_book(wb), ]


class MosComBankProfile(RequestExcelPrintFormGenerator):
    helper_class = MoscombankHelper

    def get_data(self):
        helper = self.get_helper()
        booker = helper.profile.booker
        data = {
            'L3': datetime.now().strftime('%d.%m.%Y'),
            'F5': helper.profile.full_name,
            'F6': helper.profile.short_name,
            'F8': helper.profile.get_organization_form_display(),
            'I9': helper.profile.reg_ogrn,
            'F12': helper.profile.reg_inn,
            'K12': helper.profile.reg_kpp,
            'F14': helper.profile.reg_okpo,
            'K14': helper.profile.reg_okato,
            'F15': helper.okved,
            'H16': helper.account_first.bank_account_number,
            'L16': helper.account_first.correspondent_account,
            'F17': helper.account_first.bank,
            'N17': helper.account_first.bank_bik,
            'F18': helper.getLegalAddress(),
            'F19': helper.getFactAddress(),
            'F20': helper.profile.contact_phone,
            'K21': helper.profile.contact_email,
            'F22': helper.profile.contact_name,
            'F24': helper.profile.contact_phone,
            'K24': helper.profile.contact_email,
            'K32': 'Да' if helper.profile.general_director.is_booker else 'Нет',
            'F34': helper.position_gen_dir,
            'J34': helper.position_booker,
            'F35': helper.doc_gen_dir,
            'J35': helper.doc_booker,
            'F36': '%s/%s' % (
                helper.profile.general_director.get_name,
                helper.profile.general_director.fiz_inn
            ),
            'J36': '%s/%s' % (
                helper.profile.booker.get_name,
                helper.profile.booker.fiz_inn
            ) if helper.need_booker else '',
            'F37': helper.profile.general_director.get_identity_document_name,
            'J37': helper.profile.booker.get_identity_document_name if helper.need_booker else '', # noqa
            'F38': '%s %s' % (
                helper.profile.general_director.passport.series,
                helper.profile.general_director.passport.number
            ),
            'J38': '%s %s' % (
                helper.profile.booker.passport.series,
                helper.profile.booker.passport.number) if helper.need_booker else '',
            'F39': helper.profile.general_director.passport.issued_by,
            'J39': booker.passport.issued_by if helper.need_booker else '',
            'F40': helper.profile.general_director.passport.when_issued,
            'J40': booker.passport.when_issued if helper.need_booker else '',
            'F41': helper.profile.general_director.passport.issued_code,
            'J41': booker.passport.issued_code if helper.need_booker else '',
            'F42': helper.profile.general_director.passport.date_of_birth,
            'J42': booker.passport.date_of_birth if helper.need_booker else '',
            'F43': helper.profile.general_director.passport.place_of_birth,
            'J43': booker.passport.place_of_birth if helper.need_booker else '',
            'F44': helper.profile.general_director.passport.place_of_registration,
            'J44': booker.passport.place_of_registration if helper.need_booker else '',
            'F47': helper.profile.general_director.citizenship,
            'J47': booker.citizenship if helper.need_booker else '',
            'J66': helper.profile.number_of_employees,
            'L67': helper.profile.salary_fund,
            'J67': '-',
            'J68': helper.profile.number_of_employees,
            'J69': helper.profile.number_of_employees,
            'J71': helper.experience,
            'L92': helper.has_unpaid_card,
            'L101': helper.profile.get_tax_system_display(),
            'H105': helper.request.tender.beneficiary_name,
            'F106': helper.request.tender.beneficiary_inn,
            'K106': helper.request.tender.notification_id,
            'F107': helper.targets,
            'F109': helper.request.get_contract_type_display(),
            'F111': helper.request.get_placement_way_display(),
            'E114': helper.request.protocol_number,
            'I114': helper.request.protocol_date,
            'L114': helper.request.protocol_lot_number,
            'F115': helper.request.tender.subject,
            'F116': helper.request.tender.price,
            'K117': helper.request.suggested_price_amount,
            'F118': 'Да' if helper.request.has_avans else 'Нет',
            'L118': 'Да' if Target.WARRANTY in helper.request.targets else 'Нет',
            'F119': 'Да' if helper.request.downpay else 'Нет',
            'F121': helper.request.protocol_territory,
            'J124': helper.request.required_amount,
            'K126': helper.request.interval_from,
            'K127': helper.request.interval_to,
            'J161': helper.profile.general_director.get_name,
        }

        # Добавления лицензий максимум 4
        start_license = 27
        for i, license in enumerate(helper.profile.licensessro_set.all()):
            if i > 4:
                break
            data.update({
                'B%i' % (start_license + i): license.number_license or 'нет данных',
                'F%i' % (start_license + i): license.issued_by_license or 'нет данных',
                'J%i' % (start_license + i): license.date_issue_license or 'нет данных',
                'M%i' % (start_license + i): self.date_license(license),
            })

        # Добавлени бенефициаров максимум 6
        start_beneficiar = 51
        for i, person in enumerate(helper.profile.beneficiars):
            if i > 6:
                break
            country = 'нет' if not person.resident else person.citizenship
            data.update({
                'C%i' % (start_beneficiar + i): person.get_name,
                'F%i' % (start_beneficiar + i): self.get_info_ben(person),
                'M%i' % (start_beneficiar + i): country,
                'N%i' % (start_beneficiar + i): person.share / 100,
            })

        # Добавления юр лиц максимум 5
        start_ur = 59
        for i, legal_shareholder in enumerate(helper.profile.persons_entities):
            if i > 5:
                break
            data.update({
                'C%i' % (start_ur + i): legal_shareholder.name,
                'E%i' % (start_ur + i): '%s/%s' % (
                    legal_shareholder.inn, legal_shareholder.kpp
                ),
                'H%i' % (start_ur + i): legal_shareholder.share,
                'L%i' % (start_ur + i): legal_shareholder.get_name,
            })

        # Добавление выполненых контрактов максимум 5
        start_contract = 140
        for i, contract in enumerate(helper.finished_contracts):
            if i > 5:
                break
            data.update({
                'C%i' % (start_contract + i): get_dict_to_path(
                    contract,
                    'data.customer.fullName'
                ),
                'G%i' % (start_contract + i): '.'.join(get_dict_to_path(
                    contract,
                    'data.signDate',
                ).split('-')[::-1]),
                'I%i' % (start_contract + i): contract.get('price', ''),
                'J%i' % (start_contract + i): get_dict_to_path(
                    contract,
                    'data.products.product.0.name',
                    default='',
                )
            })

        return [[*self.parse_row_and_col(k), v] for k, v in data.items()]

    @staticmethod
    def get_info_ben(ben):
        return '%s, %s, %s, %s, %s, %s' % (
            ben.fiz_inn,
            ben.passport.date_of_birth.strftime('%d.%m.%Y'),
            ben.passport.place_of_birth,
            str(ben.passport),
            ben.passport.place_of_registration,
            ben.citizenship,
        )

    @staticmethod
    def date_license(license):
        if license.is_indefinitely:
            return 'Бессрочно'
        date_end = license.date_end_license
        return date_end.strftime('%d.%m.%Y') if date_end else ''

    def generate(self):
        template_path = os.path.join(
            BASE_DIR,
            r'system_files/print_forms_templates/%s.xlsx' % self.print_form.filename
        )
        wb = self.create_book(template_path=template_path)
        ws = wb.active
        data = self.get_data()
        self.fill_sheet(ws, data)

        return [self.save_book(wb), ]


class EastExcelPrintForm(RequestExcelPrintFormGenerator):
    pass


class VoronejExcelPrintForm(RequestExcelPrintFormGenerator):
    pass


class RTBKAnketaExcelPrintForm(RequestExcelPrintFormGenerator):
    pass


class RTBKGuarantorExcelPrintForm(RequestExcelPrintFormGenerator):
    pass


class InbankConclusion(RequestExcelPrintFormGenerator):
    helper_class = InbankHelper

    def get_data(self):
        helper = self.get_helper()
        data = {
            'C5': helper.get_company_full_name(),
            'C9': helper.get_inn(),
            'G9': helper.get_reg_state_date(),
            'K9': helper.get_authorized_capital_paid(),
            'G11': helper.get_notification_id(),
            'G12': helper.get_required_amount(),
            'L12': helper.get_federal_law(),
            'G13': '[ %s ] Участие' % ('v' if helper.is_participant() else ' '),
            'I13': '[ %s ] Исполнения контракта' % (
                'v' if helper.is_execution() else ' '
            ),
            'L13': '[ %s ] Гарантийные обязательства' % (
                'v' if helper.is_warranty() else ' '
            ),
            'G14': helper.get_interval_from(),
            'I14': helper.get_interval_to(),
            'K14': helper.get_interval(),
            'G15': helper.get_tender_subject(),
            'G16': helper.get_suggested_price_amount(),
            'K16': helper.get_suggested_price_percent(),
            'G17': helper.get_beneficiar_name(),
            'D21': helper.get_executing_contracts_count(),
            'F21': (helper.get_executing_contracts_sum() or 0) / 1000000,
            'D22': helper.get_finished_contracts_count(),
            'F22': (helper.get_finished_contracts_sum() or 0) / 1000000,
            'I22': (helper.get_finished_contracts_max() or 0) / 1000000,
            'I24': helper.get_analise_period_name(),
            'C26': helper.get_last_period_name(),
            'G26': helper.get_last_period_revenues(),
            'K26': helper.get_last_period_profit(),
            'C27': helper.get_last_year_name(),
            'G27': helper.get_last_year_revenues(),
            'K27': helper.get_last_year_profit(),
            'C28': '',
            'G28': '',
            'K28': '',
            'J37': helper.finance['test_4']['value1'] or '',
            'J38': helper.finance['test_4']['value2'] or '',
            'K57': 'S' if helper.isSmallSubject else '£',
            'M57': 'S' if not helper.isSmallSubject else '£',
            'K58': 'S' if helper.isMiddleSubject else '£',
            'M58': 'S' if not helper.isMiddleSubject else '£',
            'K59': 'S' if helper.isUnknownSubject else '£',
            'M59': 'S' if not helper.isUnknownSubject else '£',
        }
        return [[*self.parse_row_and_col(k), v] for k, v in data.items()]

    def generate(self):
        template_path = os.path.join(
            BASE_DIR, r'system_files/print_forms_templates/inbank_conclusion.xlsx'
        )
        wb = self.create_book(template_path=template_path)
        ws = wb.active
        data = self.get_data()
        self.fill_sheet(ws, data)

        for el in data:
            if (el[0] == 26) and (el[1] == convert_letter_to_index('G')):
                if el[2] == '':
                    ws.merge_cells(
                        start_column=convert_letter_to_index('C'),
                        start_row=25,
                        end_column=convert_letter_to_index('F'),
                        end_row=26
                    )
                    ws['C25'].alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    ws.unmerge_cells(
                        start_column=convert_letter_to_index('G'),
                        start_row=26,
                        end_column=convert_letter_to_index('J'),
                        end_row=26
                    )
                    ws.merge_cells(
                        start_column=convert_letter_to_index('G'),
                        start_row=25,
                        end_column=convert_letter_to_index('J'),
                        end_row=26
                    )
                    ws['G25'].alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    ws.unmerge_cells(
                        start_column=convert_letter_to_index('K'),
                        start_row=26,
                        end_column=convert_letter_to_index('N'),
                        end_row=26
                    )
                    ws.merge_cells(
                        start_column=convert_letter_to_index('K'),
                        start_row=25,
                        end_column=convert_letter_to_index('N'),
                        end_row=26
                    )
                    ws['K25'].alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    # edit formula OP
                    ws.unmerge_cells('J35:J36')
                    ws['J36'].value = ws['J31'].value
                    ws.merge_cells('J35:J36')

                break

        return [self.save_book(wb), ]


class SPBConclusion(RequestExcelPrintFormGenerator):
    helper_class = SPBHelper

    def __init__(self, request, print_form, *args, **kwargs):
        super().__init__(request, print_form, *args, **kwargs)
        self.insert_dict = {}
        self.max_row = 200

    def get_data(self):
        helper = self.get_helper()

        type_product = "ГАРАНТИЯ-ТЕНДЕР-ONLINE–ГОСЗАКУПКИ"
        if Target.EXECUTION in helper.request.targets:
            type_product = 'ГАРАНТИЯ ИСПОЛНЕНИЯ ОБЯЗАТЕЛЬСТВ- ONLINE–ГОСЗАКУПКИ'

        data = {}
        booker = helper.profile.booker or helper.profile.general_director
        rating_calc = helper.client_rating_calculator.calculated_accounting_report_rating
        reverse_rating_calc = helper.reverse_client_rating_calculator
        reverse_rating_calc = reverse_rating_calc.calculated_accounting_report_rating
        principal_exp = helper.client_rating_calculator.calculated_principal_experience_rating  # noqa
        negative_facts = helper.client_rating_calculator.calculated_negative_factors_rating  # noqa

        if helper.request.prepayment:
            prepaid_expense_amount = helper.request.prepaid_expense_amount
        else:
            prepaid_expense_amount = '-'

        data['Проф суждение'] = {
            'B6': timezone.now().strftime('%d.%m.%Y'),
            'H6': helper.last_quarter.get_end_date().strftime('%d.%m.%Y'),
            'H7': helper.last_quarter.quarter,
            'E8': helper.targets,
            'E10': helper.profile.full_name,
            'E11': helper.profile.reg_inn,
            'E12': helper.profile.reg_ogrn,
            'E13': helper.profile.reg_state_date.strftime('%d.%m.%Y'),
            'E14': helper.get_main_okved(),
            'E15': 'Есть в реестре' if helper.rmsp else 'Отсутствует в реестре',
            'E16': helper.getAddressFromEGRUL(),
            'E17': helper.getFactAddress(),
            'E18': helper.profile.get_tax_system_display(),
            'E21': helper.doc_for_fact_address,
            'E22': helper.profile.authorized_capital_paid,
            'E23': helper.profile.general_director.get_name,
            'E24': booker.get_name,
            'E25': helper.profile.number_of_employees,
            'E27': helper.finished_guaranties()['total'],
            'E29': helper.request.tender.beneficiary_name,
            'E30': helper.request.tender.beneficiary_inn,
            'E31': helper.request.tender.beneficiary_ogrn,
            'E32': helper.beneficiary_okved,
            'E33': ','.join(
                [dict(Target.CHOICES)[target] for target in helper.request.targets]
            ),
            'E34': helper.request.tender.get_federal_law_display(),
            'E35': helper.request.required_amount,
            'F36': helper.request.interval_from, 'I36': helper.request.interval_to,
            'E37': helper.request.tender.notification_id,
            'E38': helper.request.tender.subject,
            'E39': prepaid_expense_amount,
            'E40': helper.request.suggested_price_amount,
            'E41': helper.request.tender.price,
            'E42': 100 - helper.request.suggested_price_percent,
            'E43': round(
                (helper.request.required_amount / helper.request.tender.price) * 100, 2
            ),
            'E44': helper.get_bank_commission_percent,
            'H50': helper.stop_factor1,
            'H51': helper.stop_factor2,
            'H53': helper.stop_factor3,
            'H55': helper.stop_factor4,
            'H57': helper.stop_factor5,
            'H58': helper.stop_factor6,
            'H59': helper.stop_factor7,
            'H60': helper.stop_factor8,
            'H61': helper.stop_factor9,
            'H63': helper.stop_factor10,
            'H64': helper.stop_factor11,
            'H65': helper.stop_factor12,
            'H66': helper.stop_factor13,
            'H67': helper.stop_factor14,
            'H68': helper.stop_factor15,
            'H69': helper.stop_factor16,
            'H70': helper.stop_factor17,
            'H71': helper.stop_factor18,
            'H73': helper.stop_factor19,
            'H74': helper.stop_factor20,
            'H76': helper.stop_factor21,
            'H77': helper.stop_factor22,
            'H78': helper.stop_factor23,
            'H79': helper.stop_factor24,
            'H80': helper.stop_factor25,
            'H81': helper.stop_factor26,
            'H82': helper.stop_factor27,
            'H83': helper.stop_factor28,
            'H85': helper.stop_factor29,
            'H87': helper.stop_factor30,
            'H88': helper.stop_factor31,
            'H90': helper.stop_factor32,
            'H91': helper.stop_factor33,
            'H93': helper.stop_factor34,
            'H94': helper.stop_factor35,
            'H95': helper.stop_factor36,
            'H96': helper.stop_factor37,
            'H97': helper.stop_factor38,
            'H98': helper.stop_factor39,
            'H99': helper.stop_factor40,
            'H100': helper.stop_factor41,
            'H101': helper.stop_factor42,
            'H103': helper.stop_factor43,
            'H104': helper.stop_factor44,
            'H105': helper.stop_factor45,
            'H106': helper.stop_factor46,
            'H107': helper.stop_factor47,
            'H108': helper.stop_factor48,
            'H109': helper.stop_factor49,

            'F115': rating_calc['formula_4']['value'][0],
            'G115': rating_calc['formula_4']['score'],
            'H115': reverse_rating_calc['formula_4']['value'][0],
            'I115': reverse_rating_calc['formula_4']['score'],
            'F116': rating_calc['formula_5']['value'][0],
            'G116': rating_calc['formula_5']['score'],
            'H116': reverse_rating_calc['formula_5']['value'][0],
            'I116': reverse_rating_calc['formula_5']['score'],
            'F117': rating_calc['formula_3']['value'][0],
            'G117': rating_calc['formula_3']['score'],
            'H117': reverse_rating_calc['formula_3']['value'][0],
            'I117': reverse_rating_calc['formula_3']['score'],
            'F118': rating_calc['formula_6']['value'][0],
            'G118': rating_calc['formula_6']['score'],
            'H118': reverse_rating_calc['formula_6']['value'][0],
            'I118': reverse_rating_calc['formula_6']['score'],
            'F119': rating_calc['formula_2']['value'][0],
            'G119': rating_calc['formula_2']['score'],
            'H119': reverse_rating_calc['formula_2']['value'][0],
            'I119': reverse_rating_calc['formula_2']['score'],
            'F120': reverse_rating_calc['formula_1']['value'][0],
            'G120': reverse_rating_calc['formula_1']['score'],
            'H120': rating_calc['formula_1']['value'][0],
            'I120': rating_calc['formula_1']['score'],
            'F121': sum(map(lambda x: x['score'], rating_calc.values())),
            'H121': sum(map(lambda x: x['score'], reverse_rating_calc.values())),
            'F122': helper.client_rating.finance_state,

            'F125': helper.net_assets_last_quarter,
            'J125': helper.net_assets_last_quarter - helper.net_assets_pre_year,
            'F126': helper.net_assets_pre_year,
            'J126': ((helper.net_assets_last_quarter / helper.net_assets_pre_year)
                     - 1) if helper.net_assets_pre_year else 0,

            'E130': helper.tender_count,
            'J130': helper.average_percent,
            'E131': helper.contract_count,
            'J131': helper.customer_count,

            'C135': len(list(filter(
                lambda x: x.status == 'E',
                helper.fz44_contracts
            ))),
            'D135': sum(map(
                lambda x: x.price,
                filter(
                    lambda x: x.status == 'E',
                    helper.fz44_contracts,
                )
            )),
            'E135': len(list(filter(
                lambda x: x.status == 'E',
                helper.fz223_contracts
            ))),
            'F135': sum(map(
                lambda x: x.price,
                filter(
                    lambda x: x.status == 'E',
                    helper.fz223_contracts,
                )
            )),
            'G135': len(list(filter(
                lambda x: x.status == 'E',
                helper.only_beneficiary
            ))),
            'H135': sum(map(
                lambda x: x.price,
                filter(
                    lambda x: x.status == 'E',
                    helper.only_beneficiary,
                )
            )),

            'C136': len(list(filter(
                lambda x: x.status == 'EC',
                helper.fz44_contracts
            ))),
            'D136': sum(map(
                lambda x: x.price,
                filter(
                    lambda x: x.status == 'EC',
                    helper.fz44_contracts,
                )
            )),
            'E136': len(list(filter(
                lambda x: x.status == 'EC',
                helper.fz223_contracts
            ))),
            'F136': sum(map(
                lambda x: x.price,
                filter(
                    lambda x: x.status == 'E',
                    helper.fz223_contracts,
                )
            )),
            'G136': len(list(filter(
                lambda x: x.status == 'EC',
                helper.only_beneficiary
            ))),
            'H136': sum(map(
                lambda x: x.price,
                filter(
                    lambda x: x.status == 'EC',
                    helper.only_beneficiary,
                )
            )),

            'C141': len(list(filter(
                lambda x: x.status == 'E',
                helper.fz44_similar_contracts
            ))),
            'D141': sum(map(
                lambda x: x.price,
                filter(
                    lambda x: x.status == 'E',
                    helper.fz44_similar_contracts,
                )
            )),
            'E141': len(list(filter(
                lambda x: x.status == 'E',
                helper.fz223_similar_contracts
            ))),
            'F141': sum(map(
                lambda x: x.price,
                filter(
                    lambda x: x.status == 'E',
                    helper.fz223_similar_contracts,
                )
            )),
            'G141': len(list(filter(
                lambda x: x.status == 'E',
                helper.only_similar_beneficiary
            ))),
            'H141': sum(map(
                lambda x: x.price,
                filter(
                    lambda x: x.status == 'E',
                    helper.only_similar_beneficiary,
                )
            )),

            'C142': len(list(filter(
                lambda x: x.status == 'EC',
                helper.fz44_similar_contracts
            ))),
            'D142': sum(map(
                lambda x: x.price or 0,
                filter(
                    lambda x: x.status == 'EC',
                    helper.fz44_similar_contracts,
                )
            )),
            'E142': len(list(filter(
                lambda x: x.status == 'EC',
                helper.fz223_similar_contracts
            ))),
            'F142': sum(map(
                lambda x: x.price,
                filter(
                    lambda x: x.status == 'EC',
                    helper.fz223_similar_contracts,
                )
            )),
            'G142': len(list(filter(
                lambda x: x.status == 'EC',
                helper.only_similar_beneficiary
            ))),
            'H142': sum(map(
                lambda x: x.price,
                filter(
                    lambda x: x.status == 'EC',
                    helper.only_similar_beneficiary,
                )
            )),

            'J145': principal_exp['formula_1']['score'],
            'J146': principal_exp['formula_2']['score'],
            'J147': principal_exp['formula_3']['score'],
            'J148': sum(map(lambda x: x['score'], principal_exp.values())),

            'J152': negative_facts['formula_1']['score'],
            'J153': negative_facts['formula_2']['score'],
            'J154': negative_facts['formula_3']['score'],
            'J155': negative_facts['formula_4']['score'],

            'J156': sum(map(lambda x: x['score'], negative_facts.values())),

            'A160': helper.client_rating.score,
            'D160': helper.client_rating.category,
            'F160': helper.client_rating.level_risk,
            'H160': helper.client_rating.finance_state,
            'A162': """Деятельность Принципала признается реальной в соответствии с п. 
            3.12.3. Положения Банка России №590-П и п.1.5 Положения Банка России 611-П 
            - на основании Решения Правления № ___ от __________.  Банковская гарантия 
            соответствует критериям ее отнесения в портфель однородных требований, 
            установленным  Приложением №6 к Положению о формировании резерва на возможные 
            потери в ПАО «Банк «Санкт-Петербург», а именно: – банковская гарантия 
            соответствует продукту «%s» (наименование продукта); – Принципал относится к 
            субъектам малого и среднего предпринимательства в соответствии с Федеральным 
            законом от 24.07.2007 г.  №209-ФЗ «О развитии малого и среднего 
            предпринимательства в РФ» (включается, если Принципал относится к субъекту 
            МСП); – сумма банковской гарантии не превышает 0,5 процента от величины 
            собственных средств ( капитала) Банка; – по условному обязательству 
            кредитного характера не выявлены индивидуальные признаки 
            обесценения """ % type_product

        }

        persons_data = [{'D': person['name'], 'J': person['share']} for person in
                        helper.profile.all_beneficiaries]
        self.insert_rows(
            20,
            persons_data,
            len(persons_data) - 1,
            'Проф суждение',
        )

        data['Приложение №2 к ПрофС'] = {
            'F6': helper.last_quarter.get_value(1100),
            'H6': helper.pre_year_quarter.get_value(1100),
            'F7': helper.last_quarter.get_value(1210),
            'H7': helper.pre_year_quarter.get_value(1210),
            'F8': helper.last_quarter.get_value(1220),
            'H8': helper.pre_year_quarter.get_value(1220),
            'F9': helper.last_quarter.get_value(1230),
            'H9': helper.pre_year_quarter.get_value(1230),
            'F12': helper.last_quarter.get_value(1240),
            'H12': helper.pre_year_quarter.get_value(1240),
            'F13': helper.last_quarter.get_value(1250),
            'H13': helper.pre_year_quarter.get_value(1250),
            'F14': helper.last_quarter.get_value(1200),
            'H14': helper.pre_year_quarter.get_value(1200),
            'F15': helper.last_quarter.get_value(1600),
            'H15': helper.pre_year_quarter.get_value(1600),
            'F17': helper.last_quarter.get_value(1300),
            'H17': helper.pre_year_quarter.get_value(1300),
            'F18': helper.last_quarter.get_value(1410),
            'H18': helper.pre_year_quarter.get_value(1410),
            'F19': helper.last_quarter.get_value(1400),
            'H19': helper.pre_year_quarter.get_value(1400),
            'F20': helper.last_quarter.get_value(1510),
            'H20': helper.pre_year_quarter.get_value(1510),
            'F21': helper.last_quarter.get_value(1520),
            'H21': helper.pre_year_quarter.get_value(1520),
            'F24': helper.last_quarter.get_value(1530),
            'H24': helper.pre_year_quarter.get_value(1530),
            'F25': helper.last_quarter.get_value(1540),
            'H25': helper.pre_year_quarter.get_value(1540),
            'F26': helper.last_quarter.get_value(1550),
            'H26': helper.pre_year_quarter.get_value(1550),
            'F27': helper.last_quarter.get_value(1500),
            'H27': helper.pre_year_quarter.get_value(1500),
            'F28': helper.last_quarter.get_value(1700),
            'H28': helper.pre_year_quarter.get_value(1700),
            'F31': helper.last_quarter.get_value(2110),
            'H31': helper.pre_year_quarter.get_value(2110),
            'F34': helper.last_quarter.get_value(2120),
            'H34': helper.pre_year_quarter.get_value(2120),
            'F35': helper.last_quarter.get_value(2200),
            'H35': helper.pre_year_quarter.get_value(2200),
            'F36': helper.last_quarter.get_value(2330),
            'H36': helper.pre_year_quarter.get_value(2330),
            'F37': helper.last_quarter.get_value(2320),
            'H37': helper.pre_year_quarter.get_value(2320),
            'F38': helper.last_quarter.get_value(2300),
            'H38': helper.pre_year_quarter.get_value(2300),
            'F39': helper.last_quarter.get_value(2400),
            'H39': helper.pre_year_quarter.get_value(2400),
            'F40': helper.last_quarter.get_value(5640),
            'H40': helper.pre_year_quarter.get_value(5640),
        }
        data['Приложение №1 к ПрофС'] = {
            'B4': helper.profile.full_name,
            'B5': helper.profile.reg_inn,
        }

        return {key: [[*self.parse_row_and_col(k), v] for k, v in el.items()] for key, el
                in data.items()}

    def action_insert(self, ws, rows_for_insert):
        for start, count in rows_for_insert:
            for _ in range(count):
                ws.insert_rows(start)
                for col in range(1, 200):
                    ws.cell(start, col)._style = ws.cell(start + 1, col)._style
                    ws.cell(start, col).number_format = ws.cell(
                        start + 1, col
                    ).number_format
                    ws.cell(start, col).value = ws.cell(start + 1, col).value

    def insert_rows(self, copy_row, data_rows, count_rows, name):
        if self.insert_dict.get(name) is None:
            self.insert_dict[name] = {
                'copy_row': [],
                'data_rows': [],
                'count_rows': [],
            }
        self.insert_dict[name]['copy_row'].append(copy_row)
        self.insert_dict[name]['data_rows'].append(data_rows)
        self.insert_dict[name]['count_rows'].append(count_rows)

    def fill_sheet(self, sheet, data, name):

        mergedcells = []
        for group in sheet.merged_cells.ranges:
            mergedcells.append(group)

        dimensions_row = []
        for i in range(1, sheet.max_row + 1):
            dimensions_row.append(sheet.row_dimensions[i].height)

        for group in mergedcells:
            sheet.unmerge_cells(str(group))

        rows_for_insert = []
        if self.insert_dict.get(name):
            rows_for_insert = list(zip(
                self.insert_dict[name]['copy_row'],
                self.insert_dict[name]['count_rows']
            ))

            for start, count, d in zip(
                    self.insert_dict[name]['copy_row'],
                    self.insert_dict[name]['count_rows'],
                    self.insert_dict[name]['data_rows']
            ):
                if count >= 1:
                    data = list(map(
                        lambda x: x if x[0] < start else [x[0] + count, x[1], x[2]],
                        data
                    ))
                for i, dict_data in enumerate(d):
                    for col, val in dict_data.items():
                        data.append([start + i, self.convert_letter_to_index(col), val])

        self.action_insert(sheet, rows_for_insert)

        for value in data:
            if len(value) > 3:
                sheet.cell(value[0], value[1]).number_format = value[3]
            sheet.cell(value[0], value[1], value[2])

        for group in mergedcells:
            for start, count in rows_for_insert:
                if group.min_row > start:
                    if count > 0:
                        group.min_row += count
                        group.max_row += count
                elif group.min_row == start:
                    if count > 0:
                        step = group.max_row - group.min_row + 1
                        for i in range(1, count + 1):
                            start_row = start + i * step
                            end_row = group.max_row + i * step
                            sheet.merge_cells(
                                start_row=start_row,
                                start_column=group.min_col,
                                end_row=end_row,
                                end_column=group.max_col
                            )
                            sheet.cell(
                                start_row,
                                group.min_col
                            )._style = sheet.cell(group.min_row, group.min_col)._style

            sheet.merge_cells(str(group))

        for i, height in enumerate(dimensions_row):
            i += 1
            for start, count in rows_for_insert:
                if i > start:
                    if count > 0:
                        i += count
                elif i == start:
                    if count > 0:
                        for i2 in range(1, count + 1):
                            sheet.row_dimensions[i + i2].height = height

            sheet.row_dimensions[i].height = height

        return sheet

    def generate(self):
        template_path = os.path.join(
            BASE_DIR, r'system_files/print_forms_templates/spb_conclusion.xlsx'
        )
        wb = self.create_book(template_path=template_path)

        for name, data in self.get_data().items():
            ws = wb.get_sheet_by_name(name)
            self.fill_sheet(ws, data, name)

        return [self.save_book(wb), ]


class SPBExtraditionDecision(SPBConclusion):

    def get_data(self):
        helper = self.get_helper()
        data = {}

        data['Решение о выдаче гарантии'] = {
            'H7': timezone.now().strftime('%d.%m.%Y'),
            'E11': helper.profile.full_name,
            'E12': helper.profile.reg_inn,
            'E13': helper.targets,
            'E14': helper.request.tender.subject,
            'E15': helper.request.request_number_in_bank,
            'E16': helper.request.tender.beneficiary_name,
            'E17': helper.request.tender.beneficiary_inn,
            'E18': helper.request.required_amount,
            'E19': helper.request.interval_from.strftime('%d.%m.%Y'),
            'E20': helper.request.interval_to.strftime('%d.%m.%Y'),
            'E21': helper.request.offer.commission,
            'E22': helper.commission_percent,
            'E23': helper.client_rating.category,
            'E24': helper.client_rating.finance_state,

        }
        requests_data = []
        bank_requests = helper.client.request_set.filter(
            bank=helper.request.bank
        ).exclude(id=helper.request.id)

        for request in bank_requests:
            if request.has_offer():
                requests_data.append({
                    'A': 'ТХ-%s, %s' % (
                        request.request_number_in_bank,
                        request.offer.contract_date,
                    ),
                    'F': request.offer.amount
                })
        data['Решение о выдаче гарантии'].update({
            'F29': sum(map(lambda x: x['F'], requests_data)),
            'F30': sum(
                map(lambda x: x['F'], requests_data)) + helper.request.required_amount,
            'A34': helper.position_signer_from_bank,
            'I34': helper.fio_signer_from_bank,
        })

        self.insert_rows(
            28,
            requests_data,
            len(requests_data) - 1,
            'Решение о выдаче гарантии'
        )
        return {
            key: [
                [*self.parse_row_and_col(k), v] for k, v in el.items()
            ] for key, el in data.items()
        }

    def generate(self):
        template_path = os.path.join(
            BASE_DIR, r'system_files/print_forms_templates/spb_extradition_decision.xlsx'
        )
        wb = self.create_book(template_path=template_path)

        for name, data in self.get_data().items():
            ws = wb.get_sheet_by_name(name)
            self.fill_sheet(ws, data, name)

        return [self.save_book(wb), ]
